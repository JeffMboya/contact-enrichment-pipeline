"""Stage 0: read the workbook by header name and pre-classify each row."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .cleaning import strip_legal_suffix
from .models import SourceRow

REQUIRED_HEADERS = ("Full name", "Address", "Company name", "Phone number")

# Some client exports name the creditor in this column.
CREDITOR_COLUMN = "Company issuing the invoice"

_FACILITY_PATTERNS = (
    "fedex drop box",
    "drop box",
    "usps",
    " scf ",
    "bmeu",
    "sectional center",
)


def _header_map(header_row) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, cell in enumerate(header_row):
        value = cell.value
        if value is not None:
            mapping[str(value).strip()] = index
    missing = [h for h in REQUIRED_HEADERS if h not in mapping]
    if missing:
        raise ValueError(
            f"Input is missing expected column(s): {', '.join(missing)}. "
            f"Found columns: {', '.join(mapping) or '(none)'}"
        )
    return mapping


def is_internal_facility(company_name: str, full_name: str) -> bool:
    haystack = f" {company_name.lower()} {full_name.lower()} "
    return any(pattern in haystack for pattern in _FACILITY_PATTERNS)


def detect_creditor_names(rows: list[SourceRow]) -> tuple[str, ...]:
    """Distinct creditor company names from the input's creditor column, with
    legal suffixes stripped for name-matching. Empty when the column is absent
    (the sample's case), so the env-configured creditor domain remains the
    fallback guard."""
    names: list[str] = []
    seen: set[str] = set()
    for row in rows:
        raw_name = (row.raw.get(CREDITOR_COLUMN) or "").strip()
        if not raw_name:
            continue
        clean, _ = strip_legal_suffix(raw_name)
        key = clean.lower()
        if clean and key not in seen:
            seen.add(key)
            names.append(clean)
    return tuple(names)


def load_rows(input_file: Path) -> list[SourceRow]:
    workbook = load_workbook(input_file, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows())
    if not rows:
        raise ValueError(f"{input_file} has no rows")

    columns = _header_map(rows[0])
    headers = [str(c.value).strip() if c.value is not None else "" for c in rows[0]]

    def cell(values, header: str) -> str:
        index = columns[header]
        value = values[index] if index < len(values) else None
        return "" if value is None else str(value).strip()

    source_rows: list[SourceRow] = []
    row_id = 0
    for excel_row in rows[1:]:
        values = [c.value for c in excel_row]
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        row_id += 1
        raw = {
            headers[i]: ("" if v is None else str(v).strip())
            for i, v in enumerate(values)
            if i < len(headers) and headers[i]
        }
        source_rows.append(
            SourceRow(
                row_id=row_id,
                raw=raw,
                full_name=cell(values, "Full name"),
                address=cell(values, "Address"),
                company_name=cell(values, "Company name"),
                phone=cell(values, "Phone number"),
            )
        )
    workbook.close()
    return source_rows
