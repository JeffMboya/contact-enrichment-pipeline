"""Stage 5: write the enriched workbook and a results.jsonl sidecar."""

from __future__ import annotations

import json
from dataclasses import asdict
from enum import Enum
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from .models import (
    ROLE_ACCOUNTS_PAYABLE,
    Contact,
    ContactType,
    EnrichmentRecord,
    NotFoundReason,
    RowStatus,
)

_APPENDED = ["Status", "Contact type", "Confidence", "Source URL", "Evidence", "Next action"]

_YELLOW = PatternFill("solid", fgColor="FFF2CC")   # Yellow fills the found-contact rows.
_GRAY = PatternFill("solid", fgColor="E0E0E0")     # Gray fills the no-contact rows.
_CONF_GREEN = PatternFill("solid", fgColor="C6EFCE")
_CONF_AMBER = PatternFill("solid", fgColor="FFEB9C")
_CONF_RED = PatternFill("solid", fgColor="FFC7CE")


def _company_label(record: EnrichmentRecord) -> str:
    resolution = record.resolution
    if resolution and resolution.domain:
        return f"{resolution.legal_name or record.parsed.clean_name} ({resolution.domain})"
    if record.parsed:
        return record.parsed.clean_name
    return record.source.company_name


def _conf_fill(value: float) -> PatternFill:
    if value >= 0.70:
        return _CONF_GREEN
    if value >= 0.40:
        return _CONF_AMBER
    return _CONF_RED


def _next_action(record: EnrichmentRecord, contact: Contact | None) -> str:
    """Plain-English instruction for a non-engineer collector: what to do with
    this row. Deterministic from status + contact type + reason — no model."""
    if record.status == RowStatus.SKIPPED:
        return "Skip: internal facility, not a debtor company."
    if contact is not None and contact.email:
        if contact.role == ROLE_ACCOUNTS_PAYABLE:
            return f"Email {contact.email} (accounts payable / billing)."
        if contact.type == ContactType.ROLE_SPECIFIC:
            return f"Email {contact.email} (role mailbox); request accounts payable."
        if contact.type == ContactType.NAMED_PERSON:
            who = f" ({contact.name})" if contact.name else ""
            return f"Email {contact.email}{who}; ask to be routed to accounts payable."
        return f"Email {contact.email} (general inbox); ask for accounts payable."
    if contact is not None and contact.type == ContactType.FORM_ONLY:
        return f"No public email. Submit the web form at {contact.source_url}."
    phone = record.source.phone
    reason = record.not_found_reason
    if record.status == RowStatus.PARTIAL:
        domain = record.resolution.domain if record.resolution else None
        site = f"{domain} " if domain else ""
        return f"Company site {site}has no public email; call {phone}." if phone else (
            f"Company site {site}has no public email; no phone on file — manual review."
        )
    if reason == NotFoundReason.NO_WEB_PRESENCE:
        return f"No web presence found; call {phone}." if phone else "No web presence found; manual review."
    if reason == NotFoundReason.AMBIGUOUS_MATCH:
        return "Manual review: a candidate was found but not confidently confirmed (see Evidence)."
    if reason == NotFoundReason.BUDGET_EXHAUSTED:
        return "Manual review: search budget reached before a confident match."
    if reason == NotFoundReason.ERROR:
        return "Manual review: processing error (see Evidence)."
    return f"Manual review; call {phone}." if phone else "Manual review."


def _added_rows(record: EnrichmentRecord) -> list[dict]:
    rows: list[dict] = []
    for contact in record.contacts:
        rows.append(
            {
                "fill": _YELLOW,
                "full_name": contact.name or "",
                "company": _company_label(record),
                "email": contact.email or "(contact form only)",
                "phone": record.source.phone,
                "status": record.status.value,
                "type": contact.type.value,
                "confidence": contact.confidence,
                "source_url": contact.source_url,
                "evidence": contact.evidence_snippet,
                "next_action": _next_action(record, contact),
            }
        )
    if not record.contacts:
        domain = record.resolution.domain if record.resolution else None
        rows.append(
            {
                "fill": _GRAY,
                "full_name": "",
                "company": _company_label(record),
                "email": "",
                "phone": record.source.phone,
                "status": record.status.value,
                "type": "",
                "confidence": 0.0,
                "source_url": domain or "",
                "evidence": record.not_found_explanation or "",
                "next_action": _next_action(record, None),
            }
        )
    return rows


def _row_is_blank(sheet, row_index: int) -> bool:
    return all(
        cell.value is None or str(cell.value).strip() == ""
        for cell in sheet[row_index]
    )


def _safe(value):
    """Neutralize spreadsheet formula injection from untrusted page content.

    openpyxl stores a string beginning with = + - @ (or a control char) as a
    live formula; a hostile contact page could land executable content in the
    workbook a human opens. Prefix such strings with a single quote.
    """
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + value
    return value


def _write_added(sheet, row_index: int, data: dict, columns: dict, base_cols: int) -> None:
    # Map by header so reordered input columns still work.
    by_header = {
        "Full name": data["full_name"],
        "Company name": data["company"],
        "Email": data["email"],
        "Phone number": data["phone"],
    }
    for col in range(1, base_cols + 1):
        cell = sheet.cell(row_index, col)
        cell.fill = data["fill"]
    for header, value in by_header.items():
        if header in columns:
            sheet.cell(row_index, columns[header], _safe(value))

    appended = [
        data["status"], data["type"], data["confidence"],
        data["source_url"], data["evidence"], data["next_action"],
    ]
    for offset, value in enumerate(appended):
        col = base_cols + 1 + offset
        cell = sheet.cell(row_index, col, _safe(value))
        cell.fill = data["fill"]
    conf_cell = sheet.cell(row_index, base_cols + 3)  # The third appended column is Confidence.
    conf_cell.fill = _conf_fill(float(data["confidence"]))
    conf_cell.number_format = "0.00"


def _to_jsonl(records: list[EnrichmentRecord]) -> str:
    def encode(obj):
        if isinstance(obj, Enum):
            return obj.value
        return str(obj)

    return "\n".join(json.dumps(asdict(r), default=encode, ensure_ascii=False) for r in records)


def write_output(
    records: list[EnrichmentRecord], input_file: Path, output_file: Path
) -> None:
    workbook = load_workbook(input_file)
    sheet = workbook.active
    base_cols = sheet.max_column

    columns = {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value is not None
    }
    for offset, header in enumerate(_APPENDED):
        sheet.cell(1, base_cols + 1 + offset, header).font = Font(bold=True)

    # Map each record to its source row by 1-based row_id.
    data_rows = [r for r in range(2, sheet.max_row + 1) if not _row_is_blank(sheet, r)]
    paired = []
    for record in records:
        idx = record.source.row_id - 1
        if not 0 <= idx < len(data_rows):
            raise ValueError(f"row_id {record.source.row_id} out of range for input")
        paired.append((record, data_rows[idx]))

    for record, sheet_row in sorted(paired, key=lambda p: p[1], reverse=True):
        added = _added_rows(record)
        sheet.insert_rows(sheet_row + 1, amount=len(added))
        for index, data in enumerate(added):
            _write_added(sheet, sheet_row + 1 + index, data, columns, base_cols)

    sheet.freeze_panes = "A2"
    _write_summary_sheet(workbook, records)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)
    output_file.with_suffix(".jsonl").write_text(_to_jsonl(records), encoding="utf-8")


def _write_summary_sheet(workbook, records: list[EnrichmentRecord]) -> None:
    """A one-glance triage tally for a collector, on a separate sheet so it never
    disturbs the row-aligned main sheet."""
    by_email = sum(1 for r in records if r.status == RowStatus.ENRICHED)
    via_form = sum(
        1 for r in records
        if r.status == RowStatus.PARTIAL
        and any(c.type == ContactType.FORM_ONLY for c in r.contacts)
    )
    manual = sum(
        1 for r in records
        if r.status == RowStatus.NOT_FOUND
        or (r.status == RowStatus.PARTIAL and not any(c.type == ContactType.FORM_ONLY for c in r.contacts))
    )
    skipped = sum(1 for r in records if r.status == RowStatus.SKIPPED)

    summary = workbook.create_sheet("Summary")
    rows = [
        ("Contact enrichment — collector worklist summary", ""),
        ("", ""),
        ("Total debtor rows", len(records)),
        ("Reachable by email (enriched)", by_email),
        ("Web form only (no email published)", via_form),
        ("Manual review needed", manual),
        ("Skipped (internal facilities)", skipped),
        ("", ""),
        ("How to read the main sheet", ""),
        ("Yellow rows", "a contact we found, inserted below its source row"),
        ("Gray rows", "no contact / skipped; see Status and Next action"),
        ("Confidence cell colour", "green ≥0.70, amber 0.40–0.69, red <0.40"),
        ("Next action column", "what to do with the row, in plain English"),
    ]
    for r, (label, value) in enumerate(rows, start=1):
        summary.cell(r, 1, _safe(label))
        summary.cell(r, 2, _safe(value))
    summary.cell(1, 1).font = Font(bold=True, size=13)
    summary.column_dimensions["A"].width = 38
    summary.column_dimensions["B"].width = 56
