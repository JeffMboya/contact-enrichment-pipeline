"""Offline tests for the key-free stages: ingestion, cleaning, extraction,
scoring, and output. No network, no API keys.

Run with:  python tests/test_offline.py

Builds a fixture workbook shaped like sample_invoices.xlsx (double-space
street/city separator, blank Email column, DUNS suffixes) so the parsing is
exercised against the real-world quirks, not a tidied-up stand-in.
"""

import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from pipeline import ingestion, output, scorer
from pipeline.cleaning import clean, is_abbreviated, query_variants, strip_duns, strip_legal_suffix
from pipeline.contact_finder import _domain_match, _email_type, extract_emails
from pipeline.resolution import parse_city_state
from pipeline.models import (
    Contact,
    ContactType,
    EnrichmentRecord,
    FullNameType,
    NotFoundReason,
    Parsed,
    Resolution,
    ResolutionStatus,
    RowStatus,
    Tier,
)

HEADERS = ["Full name", "Address", "Company name", "Email", "Phone number"]

FIXTURE_ROWS = [
    ["Joe Rich", "100 Main St  Norfolk VA 23510 USA", "DUVAL MOTORS (DUNS N° 6921522)", "", "+1 555-0100"],
    ["Invoicing", "Po Box 442  Knoxville TN 37901 USA", "POWER EQUIPMENT", "", "+1 555-0101"],
    ["Fedex Drop Box", "1 Airport Way  Memphis TN 38118 USA", "FEDEX DROP BOX (DUNS N° 119133494)", "", "+1 555-0102"],
    ["Usps Scf Pittsburgh 150", "300 Brushton Ave  Pittsburgh PA 15221 USA", "USPS SCF PITTSBURGH 150", "", "+1 555-0103"],
    ["Connie Hill", "77 River Rd  Riverside CA 92501 USA", "RIVERSIDE INFECTION CONST (DUNS N° 80646946)", "", "+1 555-0104"],
    ["Jane Doe", "9 Elm St  Concord NH 03301 USA", "LAKE CABLE LLC (DUNS N° 927410308)", "", "+1 555-0105"],
]


def make_fixture(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    for row in FIXTURE_ROWS:
        sheet.append(row)
    workbook.save(path)


def test_ingestion_and_cleaning(path: Path):
    rows = ingestion.load_rows(path)
    assert len(rows) == 6, f"expected 6 rows, got {len(rows)}"

    duval = clean(rows[0])
    assert duval.clean_name == "DUVAL MOTORS", duval.clean_name
    assert duval.duns == "6921522", duval.duns
    assert parse_city_state(rows[0].address) == ("Norfolk", "VA")

    power = clean(rows[1])
    assert power.tier == Tier.HARD, power.tier
    assert parse_city_state(rows[1].address)[0] == "Knoxville"  # PO BOX stripped

    assert ingestion.is_internal_facility(rows[2].company_name, rows[2].full_name)
    assert ingestion.is_internal_facility(rows[3].company_name, rows[3].full_name)

    riverside = clean(rows[4])
    assert is_abbreviated(riverside.clean_name), "CONST should flag truncation"
    assert "RIVERSIDE INFECTION CONSTRUCTION" in riverside.query_variants

    lake = clean(rows[5])
    assert lake.clean_name == "LAKE CABLE", lake.clean_name
    assert lake.legal_suffix == "LLC" and lake.duns == "927410308"

    # direct helpers
    assert strip_duns("ACME (DUNS N° 12)") == ("ACME", "12")
    assert strip_legal_suffix("ACME COMPANY") == ("ACME", "COMPANY")
    assert query_variants("ACME", "ACME INC") == ["ACME", "ACME INC"]
    print("ingestion + cleaning: OK")
    return rows


def test_email_extraction():
    html = (
        "<p>Accounts Payable: ap [at] duvalmotors [dot] com</p>"
        "<p>General: info@duvalmotors.com</p>"
        "<img src='logo@2x.png'>"
        "<p>tracking: noreply@sentry.io</p>"
    )
    found = {item["email"]: item for item in extract_emails(html)}
    assert "ap@duvalmotors.com" in found, "de-obfuscation failed"
    assert "info@duvalmotors.com" in found
    assert "noreply@sentry.io" not in found, "junk domain not filtered"
    assert not any(e.endswith("png") for e in found), "image artifact not filtered"
    for email, item in found.items():
        assert email in item["snippet"], "evidence snippet must contain the email"
    assert found["ap@duvalmotors.com"]["near_ap"]
    assert _email_type("ap") == ContactType.ROLE_SPECIFIC
    assert _email_type("info") == ContactType.GENERIC
    assert _email_type("jane.doe") == ContactType.NAMED_PERSON
    assert _domain_match("acme.com", "acme.com") == "exact"
    assert _domain_match("gmail.com", "acme.com") == "freemail"
    print("email extraction: OK")


def _row(company="DUVAL MOTORS"):
    return ingestion.SourceRow(1, {}, "Jane Doe", "100 Main St  Norfolk VA 23510 USA", company, "+1 555")


def test_scoring():
    row = _row()
    parsed = Parsed("DUVAL MOTORS", None, "6921522", FullNameType.PERSON, Tier.HARD, ["DUVAL MOTORS"])
    res = Resolution(ResolutionStatus.RESOLVED, "duvalmotors.com", "Duval Motors", 0.9)

    def mk(email, dm, typ, role=None, mx=True, src="https://duvalmotors.com/contact"):
        return Contact(email, None, role, typ, src, email, mx, dm)

    ap = mk("ap@duvalmotors.com", "exact", ContactType.ROLE_SPECIFIC, role="accounts payable")
    generic = mk("info@duvalmotors.com", "exact", ContactType.GENERIC)
    nomx = mk("ghost@duvalmotors.com", "exact", ContactType.ROLE_SPECIFIC, mx=False)
    scored = scorer.score_contacts([generic, ap, nomx], res, parsed, row)
    assert all(c.email != "ghost@duvalmotors.com" for c in scored), "MX gate failed"
    assert scored[0].email == "ap@duvalmotors.com", "AP should rank top"
    assert scored[0].confidence >= 0.70, scored[0].confidence

    # weak: freemail from a directory, ambiguous low-confidence resolution
    weak_res = Resolution(ResolutionStatus.AMBIGUOUS, "x.com", "X", 0.3)
    weak = mk("acme@gmail.com", "freemail", ContactType.ROLE_SPECIFIC, src="https://yelp.com/biz/x")
    assert scorer.score_contact(weak, weak_res, parsed, row) < 0.40
    print("scoring: OK")


def test_output(path: Path, rows, out_path: Path):
    res = Resolution(ResolutionStatus.RESOLVED, "duvalmotors.com", "Duval Motors", 0.9)
    contact = Contact("ap@duvalmotors.com", None, "accounts payable", ContactType.ROLE_SPECIFIC,
                      "https://duvalmotors.com/contact", "AP: ap@duvalmotors.com", True, "exact", 0.81)
    enriched = EnrichmentRecord(source=rows[0],
                                parsed=Parsed("DUVAL MOTORS", None, "6921522", FullNameType.PERSON, Tier.HARD, []),
                                resolution=res, contacts=[contact], status=RowStatus.ENRICHED)
    not_found = EnrichmentRecord(source=rows[1],
                                 parsed=Parsed("POWER EQUIPMENT", None, None, FullNameType.DEPT, Tier.HARD, []),
                                 resolution=Resolution(ResolutionStatus.UNRESOLVED, None, None, 0.0),
                                 status=RowStatus.NOT_FOUND,
                                 not_found_reason=NotFoundReason.AMBIGUOUS_MATCH,
                                 not_found_explanation="FAIL: company name too ambiguous to resolve.")
    # remaining 4 source rows need records too (row/record parity)
    fillers = [
        EnrichmentRecord(source=r, parsed=clean(r), resolution=Resolution(ResolutionStatus.UNRESOLVED, None, None, 0.0),
                         status=RowStatus.NOT_FOUND, not_found_reason=NotFoundReason.NO_WEB_PRESENCE,
                         not_found_explanation="FAIL: no web presence found.")
        for r in rows[2:]
    ]
    output.write_output([enriched, not_found, *fillers], path, out_path)

    workbook = openpyxl.load_workbook(out_path)
    sheet = workbook.active
    n = len(HEADERS)
    assert [sheet.cell(1, n + i).value for i in range(1, 7)] == \
        ["Status", "Contact type", "Confidence", "Source URL", "Evidence", "Next action"]
    email_col = HEADERS.index("Email") + 1
    assert sheet.cell(3, email_col).value == "ap@duvalmotors.com", "contact not below source"
    assert sheet.cell(3, 1).fill.fgColor.rgb.endswith("FFF2CC"), "enriched fill missing"
    assert sheet.cell(2, 1).fill.patternType is None, "source row must stay unfilled"
    conf_cell = sheet.cell(3, n + 3)
    assert conf_cell.value == 0.81 and conf_cell.number_format == "0.00"
    assert "FAIL" in str(sheet.cell(5, n + 5).value), "not-found explanation missing"
    # Next action column: a collector-readable instruction per row.
    assert "ap@duvalmotors.com" in str(sheet.cell(3, n + 6).value), "next action for enriched row"
    assert sheet.cell(5, n + 6).value, "next action for not-found row must be non-empty"
    assert sheet.cell(5, 1).fill.fgColor.rgb.endswith("E0E0E0"), "not-found gray fill missing"
    assert sheet.freeze_panes == "A2"
    # Summary sheet exists with the worklist tally.
    assert "Summary" in workbook.sheetnames, "summary sheet missing"
    assert workbook["Summary"].cell(1, 1).value.startswith("Contact enrichment")
    print("output writer: OK")


def test_creditor_guard():
    from pipeline.config import Config

    cfg = Config(None, None, Path("x"), Path("y"), Path("z"), 6, ("fedex.com",))
    assert cfg.is_creditor("fedex.com")
    assert cfg.is_creditor("www2.fedex.com")
    assert not cfg.is_creditor("duvalmotors.com")
    assert not cfg.is_creditor(None)
    print("creditor guard: OK")


def main() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        fixture = Path(tmp) / "fixture.xlsx"
        out = Path(tmp) / "enriched.xlsx"
        make_fixture(fixture)
        rows = test_ingestion_and_cleaning(fixture)
        test_email_extraction()
        test_scoring()
        test_creditor_guard()
        test_output(fixture, rows, out)
    print("\nAll offline tests passed.")


if __name__ == "__main__":
    main()
