"""Stubbed end-to-end test: resolve -> find contact -> score -> output with
canned Serper results, canned page HTML, and a canned LLM. No network, no keys.

Only the external boundaries are stubbed; everything between runs for real.

Run with:  python tests/test_stubbed_e2e.py
"""

import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from pipeline import contact_finder, ingestion, output, resolution
from pipeline.cache import Cache
from pipeline.config import Config
from pipeline.models import RowStatus

CONTACT_HTML = """
<html><body>
<h2>Contact Duval Motors — Norfolk, VA</h2>
<p>Accounts Payable inquiries: ap@duvalmotors.com</p>
<p>General: info@duvalmotors.com</p>
<p>Shipping questions: billing@fedex.com</p>
</body></html>
"""

FORM_HTML = """
<html><body>
<h2>Get in touch with Acme Widgets</h2>
<form action="/submit"><input name="message"><button>Send</button></form>
</body></html>
"""

HEADERS = ["Full name", "Address", "Company name", "Email", "Phone number"]
FIXTURE = [
    ["Joe Rich", "100 Main St  Norfolk VA 23510 USA", "DUVAL MOTORS (DUNS N° 6921522)", "", "+1 555-0100"],
    ["Fedex Drop Box", "1 Airport Way  Memphis TN 38118 USA", "FEDEX DROP BOX (DUNS N° 119133494)", "", "+1 555-0102"],
    ["Pat Lee", "5 Oak Ave  Springfield IL 62701 USA", "ACME WIDGETS", "", "+1 555-0103"],
]


def make_fixture(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    for row in FIXTURE:
        sheet.append(row)
    workbook.save(path)


def stub_serper(cache, api_key, query, num=5):
    if "ACME" in query.upper():
        return {
            "organic": [
                {"title": "Acme Widgets — Springfield IL",
                 "snippet": "Acme Widgets, 5 Oak Ave, Springfield IL 62701.",
                 "link": "https://acmewidgets.com"},
            ]
        }
    return {
        "organic": [
            {"title": "Duval Motors | Norfolk VA dealership",
             "snippet": "Duval Motors, 100 Main St, Norfolk VA 23510. Family owned.",
             "link": "https://duvalmotors.com"},
            {"title": "Duval Motors - Yelp", "snippet": "Reviews.",
             "link": "https://yelp.com/biz/duval-motors"},
        ]
    }


def stub_fetch(cache, url):
    if "duvalmotors.com" in url:
        return {"status": 200, "html": CONTACT_HTML}
    if "acmewidgets.com" in url:
        return {"status": 200, "html": FORM_HTML}
    return {"status": 404, "html": ""}


class FakeLLM:
    def json(self, model, system, user, schema):
        props = schema.get("properties", {})
        if "label" in props:
            return {"label": "person"}
        if "queries" in props:  # LLM query suggestions (hard-tier names)
            return {"queries": []}
        if "annotations" in props:  # constrained stage-3 reader
            return {"annotations": [
                {"email": "ap@duvalmotors.com", "name": "",
                 "role": "accounts payable", "department": "Finance"},
                {"email": "info@duvalmotors.com", "name": "", "role": "general",
                 "department": ""},
                # An address outside the extracted set is ignored.
                {"email": "ghost@duvalmotors.com", "name": "Phantom", "role": "", "department": ""},
            ]}
        if "result_index" in props:
            if "ACME" in user.upper():
                return {"result_index": 1, "domain": "acmewidgets.com",
                        "reasoning": "address matches", "confidence": 0.8}
            return {"result_index": 1, "domain": "duvalmotors.com",
                    "reasoning": "address in snippet matches the row", "confidence": 0.9}
        raise AssertionError(f"unexpected schema: {sorted(props)}")


def main() -> None:
    # Patch the external boundaries, then restore them afterward.
    _orig = (
        resolution.serper_search, contact_finder.serper_search,
        contact_finder.fetch_page, contact_finder._mx_valid,
    )
    resolution.serper_search = stub_serper
    contact_finder.serper_search = stub_serper
    contact_finder.fetch_page = stub_fetch
    contact_finder._mx_valid = lambda cache, domain: True

    import run as runner

    try:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            fixture = tmp_path / "fixture.xlsx"
            out = tmp_path / "enriched.xlsx"
            make_fixture(fixture)

            cfg = Config(
                serper_api_key="stub", anthropic_api_key="stub",
                input_file=fixture, output_file=out,
                cache_dir=tmp_path / "cache", max_serper_calls_per_row=6,
                creditor_domains=("fedex.com",),
            )
            cfg.cache_dir.mkdir(parents=True, exist_ok=True)
            cache = Cache(cfg.cache_dir)
            llm = FakeLLM()

            rows = ingestion.load_rows(fixture)
            records = [runner._process(row, cache, cfg, llm) for row in rows]

            duval = records[0]
            assert duval.status == RowStatus.ENRICHED, duval.status
            assert duval.resolution.domain == "duvalmotors.com", duval.resolution
            assert duval.contacts and duval.contacts[0].email == "ap@duvalmotors.com", duval.contacts
            assert duval.contacts[0].confidence >= 0.70, duval.contacts[0].confidence
            assert all("fedex.com" not in (c.email or "") for c in duval.contacts), \
                "creditor mailbox leaked into contacts"
            # A reader-named address must never reach a contact.
            assert all((c.email or "") != "ghost@duvalmotors.com" for c in duval.contacts), \
                "reader introduced an email that was not extracted from the page"

            fedex = records[1]
            assert fedex.status == RowStatus.SKIPPED, fedex.status

            acme = records[2]
            assert acme.status == RowStatus.PARTIAL, acme.status
            assert acme.contacts and acme.contacts[0].type.value == "form_only", acme.contacts
            assert acme.contacts[0].email is None
            assert "contact form" in acme.not_found_explanation, acme.not_found_explanation

            output.write_output(records, fixture, out)
            sheet = openpyxl.load_workbook(out).active
            email_col = HEADERS.index("Email") + 1
            assert sheet.cell(3, email_col).value == "ap@duvalmotors.com"
    finally:
        (resolution.serper_search, contact_finder.serper_search,
         contact_finder.fetch_page, contact_finder._mx_valid) = _orig

    print("stubbed e2e: OK")


if __name__ == "__main__":
    main()
